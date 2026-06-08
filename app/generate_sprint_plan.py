#!/usr/bin/env python3
"""Generate Sprint Plan spreadsheet (Excel/CSV).

Combines profiled tasks with assignment suggestions into a sprint planning
spreadsheet with frontend/backend responsible columns.

Usage:
    python scripts/generate_sprint_plan.py --sprint "Sprint 2025-12-22"
    python scripts/generate_sprint_plan.py --sprint "Sprint 2025-12-22" --start 2025-12-22 --end 2026-01-05
    python scripts/generate_sprint_plan.py --input data/wi_tags.json --output outputs/sprint_plan.xlsx
"""
import argparse
import json
import logging
import os
import sys
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Any, Dict, List, Optional

# Allow running from repo root
REPO_ROOT = Path(__file__).resolve().parent.parent
os.chdir(REPO_ROOT)
sys.path.insert(0, str(REPO_ROOT))

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(name)s: %(message)s"
)
logger = logging.getLogger("pm_agent.generate_sprint_plan")

OUTPUTS_DIR = REPO_ROOT / "outputs"
DATA_DIR = REPO_ROOT / "data"

# Complexity to duration mapping (in days)
COMPLEXITY_DURATION = {
    "Small": 1,
    "Medium": 3,
    "Large": 5,
    "XLarge": 8,
}

# Default status and priority
DEFAULT_STATUS = "Not Started"
DEFAULT_PRIORITY = "Medium"


def parse_args():
    parser = argparse.ArgumentParser(
        description="Generate Sprint Plan spreadsheet (Excel/CSV)"
    )
    parser.add_argument(
        "--sprint", "-s",
        type=str,
        required=True,
        help="Sprint name (e.g., 'Sprint 2025-12-22')"
    )
    parser.add_argument(
        "--start",
        type=str,
        default=None,
        help="Sprint start date (YYYY-MM-DD). Defaults to today."
    )
    parser.add_argument(
        "--end",
        type=str,
        default=None,
        help="Sprint end date (YYYY-MM-DD). Defaults to start + 14 days."
    )
    parser.add_argument(
        "--input", "-i",
        type=str,
        default=None,
        help="Path to profiled tasks JSON. Default: data/wi_tags.json"
    )
    parser.add_argument(
        "--assignments", "-a",
        type=str,
        default=None,
        help="Path to assignment suggestions JSON. Default: data/role_assignment_suggestions.json"
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        default=None,
        help="Output path (without extension). Default: outputs/sprint_plan_<ts>"
    )
    parser.add_argument(
        "--format",
        type=str,
        choices=["excel", "csv", "both"],
        default="both",
        help="Output format: excel, csv, or both (default: both)"
    )
    return parser.parse_args()


def load_profiled_tasks(input_path: Optional[Path] = None) -> List[Dict]:
    """Load profiled tasks from JSON file."""
    if input_path and Path(input_path).exists():
        with open(input_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        if isinstance(data, list):
            return data
        return data.get("items", [])
    
    # Default: data/wi_tags.json
    tags_file = DATA_DIR / "wi_tags.json"
    if tags_file.exists():
        with tags_file.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return data.get("items", list(data.values()) if isinstance(data, dict) else [])
    
    return []


def load_role_assignments(assignments_path: Optional[Path] = None) -> Dict[int, Dict]:
    """Load role-based assignment suggestions, keyed by WI ID."""
    if assignments_path and Path(assignments_path).exists():
        path = Path(assignments_path)
    else:
        path = DATA_DIR / "role_assignment_suggestions.json"
    
    if not path.exists():
        # Fall back to regular assignment suggestions
        path = DATA_DIR / "assignment_suggestions.json"
    
    if not path.exists():
        return {}
    
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        
        # Build lookup by WI ID
        lookup = {}
        for item in data.get("suggestions", []):
            wi_id = item.get("wi_id")
            if wi_id:
                lookup[wi_id] = item
        return lookup
    except Exception as e:
        logger.warning("Failed to load assignments: %s", e)
        return {}


def load_manual_overrides() -> Dict[int, Dict]:
    """Load any manual overrides from data/sprint_plan_overrides.json."""
    overrides_file = DATA_DIR / "sprint_plan_overrides.json"
    if not overrides_file.exists():
        return {}
    
    try:
        with overrides_file.open("r", encoding="utf-8") as f:
            data = json.load(f)
        return {int(k): v for k, v in data.items()}
    except Exception:
        return {}


def save_manual_overrides(overrides: Dict[int, Dict]):
    """Save manual overrides."""
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    overrides_file = DATA_DIR / "sprint_plan_overrides.json"
    
    with overrides_file.open("w", encoding="utf-8") as f:
        json.dump(overrides, f, indent=2)


def get_top_suggestion(suggestions: List[Dict], fallback: str = "") -> str:
    """Get the top developer suggestion username."""
    if not suggestions:
        return fallback
    return suggestions[0].get("developer", fallback).split("@")[0]


def get_suggestion_metadata(suggestions: List[Dict]) -> str:
    """Build metadata string for hidden column."""
    if not suggestions:
        return ""
    
    parts = []
    for i, sug in enumerate(suggestions[:3], 1):
        dev = sug.get("developer", "").split("@")[0]
        score = sug.get("score", 0)
        breakdown = sug.get("breakdown", {})
        skill = breakdown.get("skill_match", 0)
        fam = breakdown.get("familiarity", 0)
        parts.append(f"{i}. {dev} (score={score:.0%}, skill={skill:.0%}, fam={fam:.0%})")
    
    return "; ".join(parts)


def estimate_dates(
    start_date: datetime,
    complexity: str,
    task_index: int,
) -> tuple:
    """Estimate start and end dates based on complexity and ordering."""
    duration_days = COMPLEXITY_DURATION.get(complexity, 3)
    
    # Stagger tasks slightly
    task_start = start_date + timedelta(days=task_index % 5)
    task_end = task_start + timedelta(days=duration_days)
    
    return task_start.strftime("%Y-%m-%d"), task_end.strftime("%Y-%m-%d"), duration_days


def build_sprint_plan_rows(
    tasks: List[Dict],
    assignments: Dict[int, Dict],
    overrides: Dict[int, Dict],
    sprint_name: str,
    sprint_start: datetime,
    sprint_end: datetime,
) -> List[Dict]:
    """Build rows for the sprint plan spreadsheet."""
    rows = []
    
    for idx, task in enumerate(tasks):
        wi_id = task.get("id")
        title = task.get("title", "")
        area_path = task.get("area_path", "")
        complexity = task.get("complexity", "Medium")
        work_item_type = task.get("work_item_type", "User Story")
        
        # Get assignment suggestions
        assignment = assignments.get(wi_id, {})
        fe_suggestions = assignment.get("frontend_suggestions", assignment.get("suggestions", []))
        be_suggestions = assignment.get("backend_suggestions", [])
        
        # Apply manual overrides
        override = overrides.get(wi_id, {})
        
        # Estimate dates
        est_start, est_end, duration = estimate_dates(sprint_start, complexity, idx)
        
        # Build row
        row = {
            "Sprint": sprint_name,
            "Feature / User Story": title,
            "Task Name": f"WI-{wi_id}",
            "WI ID": wi_id,
            "Work Item Type": work_item_type,
            "Area Path": area_path,
            "Responsible - Frontend": override.get("frontend") or get_top_suggestion(fe_suggestions, "TBD"),
            "Frontend Suggestions": get_suggestion_metadata(fe_suggestions),
            "Responsible - Backend": override.get("backend") or get_top_suggestion(be_suggestions, "TBD"),
            "Backend Suggestions": get_suggestion_metadata(be_suggestions),
            "Start Date": override.get("start_date") or est_start,
            "End Date": override.get("end_date") or est_end,
            "Duration (days)": duration,
            "Status": override.get("status") or DEFAULT_STATUS,
            "Priority": override.get("priority") or DEFAULT_PRIORITY,
            "Complexity": complexity,
            "Comments": override.get("comments") or "",
        }
        
        rows.append(row)
    
    return rows


def generate_excel(rows: List[Dict], output_path: Path, sprint_name: str):
    """Generate Excel file with formatting."""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed. Run: pip install openpyxl")
        return None
    
    import pandas as pd
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sprint Plan"
    
    # Define columns (visible)
    visible_columns = [
        "Sprint", "Feature / User Story", "Task Name",
        "Responsible - Frontend", "Responsible - Backend",
        "Start Date", "End Date", "Duration (days)",
        "Status", "Priority", "Comments"
    ]
    
    # Hidden/metadata columns
    hidden_columns = [
        "WI ID", "Work Item Type", "Area Path", "Complexity",
        "Frontend Suggestions", "Backend Suggestions"
    ]
    
    all_columns = visible_columns + hidden_columns
    
    # Write header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    for col_idx, col_name in enumerate(all_columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Write data rows
    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, col_name in enumerate(all_columns, 1):
            value = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Add duration formula for Excel
            if col_name == "Duration (days)":
                # Excel formula: =G2-F2 (End Date - Start Date)
                end_col = get_column_letter(all_columns.index("End Date") + 1)
                start_col = get_column_letter(all_columns.index("Start Date") + 1)
                cell.value = f"={end_col}{row_idx}-{start_col}{row_idx}"
            
            # Color code status
            if col_name == "Status":
                if value == "Done":
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                elif value == "In Progress":
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif value == "Blocked":
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            
            # Color code priority
            if col_name == "Priority":
                if value == "Critical":
                    cell.font = Font(bold=True, color="9C0006")
                elif value == "High":
                    cell.font = Font(color="9C5700")
    
    # Set column widths
    column_widths = {
        "Sprint": 18,
        "Feature / User Story": 50,
        "Task Name": 12,
        "Responsible - Frontend": 20,
        "Responsible - Backend": 20,
        "Start Date": 12,
        "End Date": 12,
        "Duration (days)": 10,
        "Status": 12,
        "Priority": 10,
        "Comments": 30,
        "WI ID": 8,
        "Work Item Type": 12,
        "Area Path": 25,
        "Complexity": 10,
        "Frontend Suggestions": 50,
        "Backend Suggestions": 50,
    }
    
    for col_idx, col_name in enumerate(all_columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = column_widths.get(col_name, 15)
    
    # Hide metadata columns
    for col_name in hidden_columns:
        col_idx = all_columns.index(col_name) + 1
        ws.column_dimensions[get_column_letter(col_idx)].hidden = True
    
    # Freeze panes (header row and first 3 columns)
    ws.freeze_panes = "D2"
    
    # Add borders
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in ws.iter_rows(min_row=1, max_row=len(rows) + 1, max_col=len(all_columns)):
        for cell in row:
            cell.border = thin_border
    
    # Save
    wb.save(output_path)
    logger.info("Excel file saved: %s", output_path)
    return output_path


def generate_csv(rows: List[Dict], output_path: Path):
    """Generate CSV file."""
    import pandas as pd
    
    # All columns for CSV (no hidden columns, just include all)
    columns = [
        "Sprint", "Feature / User Story", "Task Name", "WI ID",
        "Responsible - Frontend", "Frontend Suggestions",
        "Responsible - Backend", "Backend Suggestions",
        "Start Date", "End Date", "Duration (days)",
        "Status", "Priority", "Complexity", "Comments"
    ]
    
    df = pd.DataFrame(rows)
    
    # Reorder and select columns
    available_cols = [c for c in columns if c in df.columns]
    df = df[available_cols]
    
    df.to_csv(output_path, index=False)
    logger.info("CSV file saved: %s", output_path)
    return output_path


def main():
    args = parse_args()
    
    # Parse dates
    if args.start:
        sprint_start = datetime.strptime(args.start, "%Y-%m-%d")
    else:
        sprint_start = datetime.now()
    
    if args.end:
        sprint_end = datetime.strptime(args.end, "%Y-%m-%d")
    else:
        sprint_end = sprint_start + timedelta(days=14)
    
    logger.info("=" * 60)
    logger.info("GENERATING SPRINT PLAN")
    logger.info("=" * 60)
    logger.info("Sprint: %s", args.sprint)
    logger.info("Start: %s", sprint_start.strftime("%Y-%m-%d"))
    logger.info("End: %s", sprint_end.strftime("%Y-%m-%d"))
    
    # Load data
    tasks = load_profiled_tasks(args.input)
    if not tasks:
        logger.error("No profiled tasks found. Run profile_upcoming_tasks.py first.")
        return 1
    
    logger.info("Loaded %d profiled tasks", len(tasks))
    
    # Load assignments
    assignments = load_role_assignments(args.assignments)
    if not assignments:
        # Generate role-based assignments if not found
        logger.info("No role-based assignments found, generating...")
        from utilities.assignment import run_role_based_assignment_pipeline
        suggestions = run_role_based_assignment_pipeline(profiled_tasks=tasks)
        assignments = {s["wi_id"]: s for s in suggestions}
    
    logger.info("Loaded assignments for %d WIs", len(assignments))
    
    # Load manual overrides
    overrides = load_manual_overrides()
    if overrides:
        logger.info("Loaded %d manual overrides", len(overrides))
    
    # Build rows
    rows = build_sprint_plan_rows(
        tasks, assignments, overrides,
        args.sprint, sprint_start, sprint_end
    )
    
    # Determine output path
    ts = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")
    if args.output:
        base_path = Path(args.output)
    else:
        OUTPUTS_DIR.mkdir(parents=True, exist_ok=True)
        base_path = OUTPUTS_DIR / f"sprint_plan_{ts}"
    
    outputs = []
    
    # Generate outputs
    if args.format in ("excel", "both"):
        excel_path = base_path.with_suffix(".xlsx")
        result = generate_excel(rows, excel_path, args.sprint)
        if result:
            outputs.append(str(excel_path))
    
    if args.format in ("csv", "both"):
        csv_path = base_path.with_suffix(".csv")
        generate_csv(rows, csv_path)
        outputs.append(str(csv_path))
    
    # Summary
    print("\n" + "=" * 60)
    print("SPRINT PLAN GENERATION COMPLETE")
    print("=" * 60)
    print(f"Sprint: {args.sprint}")
    print(f"Tasks planned: {len(rows)}")
    print(f"Output files:")
    for out in outputs:
        print(f"  - {out}")
    
    # Show sample rows
    print("\n" + "-" * 60)
    print("SAMPLE ROWS (first 5)")
    print("-" * 60)
    
    for row in rows[:5]:
        print(f"\n[{row['WI ID']}] {row['Feature / User Story'][:50]}...")
        print(f"  Frontend: {row['Responsible - Frontend']}")
        print(f"  Backend: {row['Responsible - Backend']}")
        print(f"  Dates: {row['Start Date']} → {row['End Date']}")
        print(f"  Status: {row['Status']} | Priority: {row['Priority']}")
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
